import socket
import openpyxl
import threading
from time import sleep


# HOST E PORTA PADRÃO DO SERVIDOR
HOST = 'localhost'  # 127.0.0.1
PORT = 50000


def handle_client(socket_client, client_address):
    print(f'Conectado em {client_address}')  # cliente conectado

    # IMPORTA A BASE DE DADOS
    database = openpyxl.load_workbook('database.xlsx')
    table = database.active

    # CRIA UMA LISTA COM TODOS OS ANIMAIS DISPONÍVEIS
    animals = [table.cell(row=row, column=1).value for row in range(
        2, table.max_row + 1) if table.cell(row=row, column=1).value]

    # CRIA UMA LISTA COM TODOS OS ATRIBUTOS DISPONÍVEIS
    atributes = [table.cell(row=1, column=column).value for column in range(
        2, table.max_column + 1) if table.cell(row=1, column=column).value]

    # CRIA UMA MATRIZ COM TODOS OS VALORES BOOLEANOS
    values_matrix = []
    for row in range(2, table.max_row + 1):
        values = []
        for column in range(2, table.max_column + 1):
            values.append(table.cell(row=row, column=column).value)

        values_matrix.append(values)

    animalsIdx = range(len(values_matrix))
    attributesIdx = range(len(values_matrix[0]))

    # EXECUTA A ÀRVORE DE DECISÃO
    for attributeIdx in attributesIdx:
        attributeFactibility = False
        for animalIdx in animalsIdx:
            if values_matrix[animalIdx][attributeIdx]:
                attributeFactibility = True
                break
        if not attributeFactibility:
            continue
        mensage = f'O animal que você está pensando {atributes[attributeIdx]}? (s/n): '
        socket_client.sendall(mensage.encode())
        r = socket_client.recv(4096).decode()
        if r == 's':
            search = None
        else:
            search = 1
        animalsIdxAux = [animal for animal in animalsIdx]
        for animalIdx in animalsIdx:
            if values_matrix[animalIdx][attributeIdx] == search:
                animalsIdxAux.remove(animalIdx)
        if len(animalsIdxAux) == 1:
            # MOSTRA OS RESULTADO
            mensage = f'Eu acho que o animal que você está pensando é um(a) {animals[animalsIdxAux[0]]}.\nEu acertei? (s/n): '
            socket_client.sendall(mensage.encode())
            data = socket_client.recv(4096) # resposta do cliente
            if data.decode() == 's':
                mensage = 'Hehehehehe. Eu nunca erro!! :)'
                socket_client.sendall(mensage.encode())
            else:
                mensage = 'Sério? Desculpa, não sei o que aconteceu. :('
                socket_client.sendall(mensage.encode())
            socket_client.close() 
            animalsIdx = animalsIdxAux
            break
        animalsIdx = animalsIdxAux

    if len(animalsIdx) > 1:
        print(f'{", ".join([animals[animalIdx] for animalIdx in animalsIdx ])}')


# criando um socket do tipo TCP
with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as socket_server:
    # associa um socket a um endereço e uma porta especifica
    socket_server.bind((HOST, PORT))
    socket_server.listen()  # começa a "escutar" as conexões
    print('Aguardando conexão de um cliente...')

    while True:  # loop que fica sempre procurando conexões
        # aceitando a conexão do client
        socket_client, client_address = socket_server.accept()
        print('\nNovo cliente conectado. Inicializando jogo...')
        client_thread = threading.Thread(target=handle_client, args=(socket_client, client_address)) # cria uma thread para cada cliente conectado
        client_thread.start() # iniciando a thread do client
