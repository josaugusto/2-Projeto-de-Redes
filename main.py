import openpyxl


# IMPORTA A BASE DE DADOS
database = openpyxl.load_workbook('database.xlsx')
table = database.active

# CRIA UMA LISTA COM TODOS OS ANIMAIS DISPONÍVEIS
animals = [table.cell(row=row, column=1).value for row in range(2, table.max_row + 1) if table.cell(row=row, column=1).value]

# CRIA UMA LISTA COM TODOS OS ATRIBUTOS DISPONÍVEIS
atributes = [table.cell(row=1, column=column).value for column in range(2, table.max_column + 1) if table.cell(row=1, column=column).value]

# CRIA UMA MATRIZ COM TODOS OS VALORES BOOLEANOS
values_matrix = []
for row in range(2, table.max_row + 1):
    values = []
    for column in range(2, table.max_column + 1):
        values.append(table.cell(row=row, column=column).value)

    values_matrix.append(values)    

animalsIdx = range(len(values_matrix))
attributesIdx = range(len(values_matrix[0]))


# EXECUTA A ÀRVORE DE DECISÃO
for attributeIdx in attributesIdx:
    attributeFactibility = False
    for animalIdx in animalsIdx:
        if values_matrix[animalIdx][attributeIdx]:
            attributeFactibility = True
            break
    if not attributeFactibility:
        continue
    r = input(f'O animal que você está pensando {atributes[attributeIdx]}? (s/n): ')
    if r == 's':
        search = None
    else:
        search = 1 
    animalsIdxAux = [animal for animal in animalsIdx]       
    for animalIdx in animalsIdx:
        if values_matrix[animalIdx][attributeIdx] == search:
            animalsIdxAux.remove(animalIdx)
    if len(animalsIdxAux) == 1:
        
        # MOSTRA OS RESULTADO
        print(f'Eu acho que o animal que você está pensando é um(a) {animals[animalsIdxAux[0]]}.')
        if input('Eu acertei? (s/n): ') == 's':
            print('Hehehehehe. Eu nunca erro!! :)')

        else:
            print('Sério? Desculpa, não sei o que aconteceu. :(')
        animalsIdx = animalsIdxAux
        break
    animalsIdx = animalsIdxAux

if len(animalsIdx) > 1:
    print(f'{", ".join([animals[animalIdx] for animalIdx in animalsIdx ])}')